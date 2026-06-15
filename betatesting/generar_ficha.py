#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
generar_ficha.py
================
Generador de la plantilla Excel ``Ficha_Betatester.xlsx`` para el betatesting
de **Argos2**.

Uso:
    python betatesting/generar_ficha.py

El script crea (o sobrescribe) ``Ficha_Betatester.xlsx`` en la RAÍZ del proyecto
(un nivel por encima de esta carpeta).

Requisitos
----------
* ``openpyxl`` (``pip install openpyxl``).

NOTA: openpyxl **NO** es una dependencia de runtime de la aplicación Argos2;
solo se usa aquí para construir la plantilla. Por eso **no** debe añadirse a
``Backend/requirements.txt``. El ``.xlsx`` resultante es un archivo estático
que los betatesters abren con Excel/LibreOffice sin necesidad de Python.

Estructura del libro generado (4 hojas, la última oculta):
1. "Instrucciones"        -> guía de uso para el betatester.
2. "Reportes"             -> ficha rellenable principal (13 columnas).
3. "Checklist de Pruebas" -> lista de funcionalidades a marcar.
4. "Listas" (oculta)      -> valores de los desplegables (DataValidation).
"""

from __future__ import annotations

import os
import sys
from datetime import date

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# La consola de Windows usa cp1252 por defecto; forzamos UTF-8 para que las
# tildes y los emojis se impriman sin UnicodeEncodeError.
_reconfigure = getattr(sys.stdout, "reconfigure", None)
if _reconfigure is not None:
    try:
        _reconfigure(encoding="utf-8")
    except Exception:
        pass

# ----------------------------------------------------------------------
# Ruta de salida: raíz del proyecto (padre de esta carpeta).
# ----------------------------------------------------------------------
AQUI = os.path.dirname(os.path.abspath(__file__))
RAIZ_PROYECTO = os.path.dirname(AQUI)
SALIDA = os.path.join(RAIZ_PROYECTO, "Ficha_Betatester.xlsx")

# ----------------------------------------------------------------------
# Valores de los desplegables
# ----------------------------------------------------------------------
MODULOS = [
    "Autenticación",
    "Correo",
    "Cámaras (Descubrimiento/Registro)",
    "Monitoreo en Vivo",
    "Captura/Galería",
    "Visión (Cloud/Local)",
    "Selector de Visión",
    "Panel de Ajustes",
    "Dashboard/Pestañas",
    "Administración (Usuarios)",
    "Administración (Cámaras)",
    "PWA/Instalación",
    "Seguridad/Rate Limiting",
    "Rendimiento",
    "UI/UX",
    "Otro",
]

TIPOS = [
    "Bug funcional",
    "Error visual/CSS",
    "Bug de rendimiento/lentitud",
    "Fallo de seguridad",
    "Sugerencia de mejora",
    "Duda/Consulta",
    "Documentación",
    "Otro",
]

SEVERIDADES = [
    "Crítica (bloquea uso)",
    "Alta (función clave rota)",
    "Media (workaround existe)",
    "Baja (cosmético/menor)",
    "Informativa",
]

ESTADOS = ["Nuevo", "En revisión", "Confirmado", "En progreso", "Resuelto",
           "No reproduce", "Cerrado"]

# Checklist
OPC_PROBADA = ["Sí", "No", "No aplica"]
OPC_RESULTADO = ["OK", "Con errores", "No pude probarla"]

# ----------------------------------------------------------------------
# Encabezados de las hojas rellenables
# ----------------------------------------------------------------------
REPORTES_HEADERS = [
    "ID Reporte",                # A
    "Fecha",                     # B
    "Betatester",                # C
    "Módulo",                    # D
    "Tipo",                      # E
    "Severidad",                 # F
    "Navegador/SO",              # G
    "Pasos para reproducir",     # H
    "Resultado esperado",        # I
    "Resultado actual",          # J
    "Evidencia (archivo/ruta)",  # K
    "Estado",                    # L
    "Notas/Resolución",          # M
]

CHECKLIST_HEADERS = [
    "#",                          # A
    "Funcionalidad",             # B
    "¿Probada?",                 # C
    "Resultado",                 # D
    "ID Reporte relacionado",    # E
    "Comentarios",               # F
]

FUNCIONALIDADES = [
    "Instalación (install.bat / install.sh)",
    "Registro de cuenta",
    "Verificación por correo (código de 6 dígitos)",
    "Inicio de sesión (login)",
    "Cierre de sesión (logout)",
    "Recuperación de contraseña",
    "Restablecimiento de contraseña",
    "Descubrimiento de cámaras USB",
    "Registro de cámara IP / ESP32-CAM",
    "Monitoreo en vivo (stream MJPEG)",
    "Reconexión automática ante caída de cámara",
    "Captura de fotos",
    "Galería de capturas (FIFO)",
    "Descarga de capturas",
    "Selector de visión por cámara (Off/Cloud/Local)",
    "Visión Cloud (Roboflow) — bounding boxes",
    "Visión Local",
    "Persistencia de selección de visión (recarga)",
    "Panel de Ajustes (API key enmascarada, guardar, test conexión)",
    "Dashboard con pestañas (visibilidad por rol)",
    "Pantalla completa de cámara",
    "Panel de administración — gestión de usuarios (rol/estado)",
    "Protección anti-self (no auto-modificarse/auto-eliminarse)",
    "Escaneo de red ESP32-CAM",
    "Reiniciar/eliminar cámara desde admin",
    "Rate limiting (bloqueo tras reintentos)",
    "PWA — instalación",
    "PWA — modo offline",
    "Endpoint /health",
    "Endpoint /api (documentación)",
]

# ----------------------------------------------------------------------
# Estilos comunes
# ----------------------------------------------------------------------
COLOR_HEADER = "1F4E78"        # azul oscuro
COLOR_HEADER_CHECK = "375623"  # verde oscuro
COLOR_TITULO = "1F4E78"
COLOR_EJEMPLO = "FFE699"       # ámbar claro para filas de ejemplo
COLOR_NOTA = "FCE4D6"          # salmón claro
COLOR_BANDA = "F2F2F2"         # gris muy claro (bandas)

THIN = Side(style="thin", color="BFBFBF")
BORDE_CELDA = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FUENTE_HEADER = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
FUENTE_TITULO = Font(name="Calibri", size=18, bold=True, color="1F4E78")
FUENTE_SUBTITULO = Font(name="Calibri", size=13, bold=True, color="1F4E78")
FUENTE_NORMAL = Font(name="Calibri", size=11)
FUENTE_NEGRITA = Font(name="Calibri", size=11, bold=True)
FUENTE_EJEMPLO = Font(name="Calibri", size=11, italic=True, color="7F6000")

ALIGN_HEADER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_TEXTO = Alignment(horizontal="left", vertical="top", wrap_text=True)
ALIGN_CENTRO = Alignment(horizontal="center", vertical="center", wrap_text=True)


def letra(n):
    """Número de columna -> letra ('A', 'B', ...)."""
    return get_column_letter(n)


def llenar_encabezados(ws, headers, color, fila=1):
    """Escribe la fila de encabezados con estilo."""
    fill = PatternFill("solid", fgColor=color)
    for col_idx, titulo in enumerate(headers, start=1):
        celda = ws.cell(row=fila, column=col_idx, value=titulo)
        celda.fill = fill
        celda.font = FUENTE_HEADER
        celda.alignment = ALIGN_HEADER
        celda.border = BORDE_CELDA
    ws.row_dimensions[fila].height = 32


def anadir_lista(ws, columna_dest, fila_ini, fila_fin, hoja_lista, col_lista,
                 cantidad):
    """Crea y aplica una validación de lista que apunta a un rango de 'Listas'.

    columna_dest : columna destino en la hoja actual (p. ej. 'D').
    col_lista    : columna de la hoja 'Listas' con los valores ('A'..'F').
    cantidad     : nº de valores (filas) en esa columna de 'Listas'.
    """
    rango_lista = f"{hoja_lista}!${col_lista}$1:${col_lista}${cantidad}"
    dv = DataValidation(type="list", formula1=rango_lista, allow_blank=True)
    dv.error = "Por favor, selecciona un valor de la lista desplegable."
    dv.errorTitle = "Valor no válido"
    dv.prompt = "Selecciona una opción de la lista."
    dv.promptTitle = "Elige una opción"
    ws.add_data_validation(dv)
    dv.add(f"{columna_dest}{fila_ini}:{columna_dest}{fila_fin}")


# ======================================================================
# HOJA 1: Instrucciones
# ======================================================================
def crear_hoja_instrucciones(wb: Workbook):
    ws = wb.create_sheet("Instrucciones")
    ws.sheet_properties.tabColor = COLOR_TITULO
    ws.sheet_view.showGridLines = False

    # Anchos cómodos para lectura
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 70
    ws.column_dimensions["D"].width = 3

    fila = 2
    # Título
    ws.merge_cells(start_row=fila, start_column=2, end_row=fila, end_column=3)
    c = ws.cell(row=fila, column=2,
                value="📋 FICHA DE REPORTE — Betatesting Argos2")
    c.font = FUENTE_TITULO
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[fila].height = 30
    fila += 2

    # Párrafo explicativo
    ws.merge_cells(start_row=fila, start_column=2, end_row=fila, end_column=3)
    c = ws.cell(
        row=fila, column=2,
        value=("Esta ficha te permite reportar errores, observaciones o "
               "sugerencias que encuentres al probar Argos2. Rellena una fila "
               "por cada reporte en la hoja 'Reportes'."))
    c.font = FUENTE_NORMAL
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.row_dimensions[fila].height = 45
    fila += 2

    # Pasos
    c = ws.cell(row=fila, column=2, value="Pasos para usarla:")
    c.font = FUENTE_SUBTITULO
    fila += 1
    pasos = [
        'Ve a la hoja "Reportes".',
        ("Rellena una fila por cada problema u observación que encuentres. "
         "Las columnas Módulo, Tipo, Severidad y Estado tienen menús "
         "desplegables (solo selecciona una opción)."),
        ('Sé claro y específico en "Pasos para reproducir", '
         '"Resultado esperado" y "Resultado actual".'),
        ("Si tienes una captura de pantalla o video, guarda el archivo y pon "
         'su nombre o ruta en la columna "Evidencia".'),
        ("Cuando termines, guarda este archivo y envíalo por correo a "
         "sqprpject@gmail.com."),
    ]
    for i, paso in enumerate(pasos, start=1):
        ws.merge_cells(start_row=fila, start_column=2,
                       end_row=fila, end_column=3)
        c = ws.cell(row=fila, column=2, value=f"{i}.  {paso}")
        c.font = FUENTE_NORMAL
        c.alignment = Alignment(horizontal="left", vertical="top",
                                wrap_text=True)
        ws.row_dimensions[fila].height = 32
        fila += 1

    fila += 1
    # Nota
    ws.merge_cells(start_row=fila, start_column=2,
                   end_row=fila, end_column=3)
    c = ws.cell(
        row=fila, column=2,
        value=("ℹ️ Nota: Si no sabes cuál módulo elegir, selecciona 'Otro'. "
               "Lo importante es que describas bien el problema."))
    c.font = Font(name="Calibri", size=11, italic=True, color="7F6000")
    c.fill = PatternFill("solid", fgColor=COLOR_NOTA)
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.row_dimensions[fila].height = 32
    fila += 2

    # ---- Tabla de Severidad ----
    c = ws.cell(row=fila, column=2, value="Niveles de Severidad")
    c.font = FUENTE_SUBTITULO
    fila += 1
    fila = tabla_referencia(
        ws, fila,
        ["Nivel", "Significado"],
        [
            ("Crítica (bloquea uso)",
             "Impide usar la aplicación por completo. No hay forma de continuar."),
            ("Alta (función clave rota)",
             "Una función principal no funciona, pero la app no se cae entera."),
            ("Media (workaround existe)",
             "Hay un error, pero existe una forma alternativa de lograr la tarea."),
            ("Baja (cosmético/menor)",
             "Detalle visual o menor que no afecta a la funcionalidad."),
            ("Informativa",
             "Comentario, sugerencia u observación; no es un error como tal."),
        ],
    )
    fila += 2

    # ---- Tabla de Tipos ----
    c = ws.cell(row=fila, column=2, value="Tipos de reporte")
    c.font = FUENTE_SUBTITULO
    fila += 1
    tabla_referencia(
        ws, fila,
        ["Tipo", "Significado"],
        [
            ("Bug funcional", "Algo no funciona como debería."),
            ("Error visual/CSS",
             "Fallo de diseño: desalineación, colores, superposición, etc."),
            ("Bug de rendimiento/lentitud",
             "Va lento, se cuelga o consume muchos recursos."),
            ("Fallo de seguridad",
             "Posible brecha, acceso sin permiso o datos expuestos."),
            ("Sugerencia de mejora", "Idea para mejorar la aplicación."),
            ("Duda/Consulta", "Pregunta sobre cómo funciona algo."),
            ("Documentación", "Error o ausencia en la ayuda o el manual."),
            ("Otro", "Cualquier cosa que no encaje en las categorías anteriores."),
        ],
    )

    return ws


def tabla_referencia(ws, fila, titulos, filas_datos):
    """Dibuja una tabla de 2 columnas con cabecera y devuelve la siguiente fila libre."""
    # Cabecera
    for col, t in enumerate(titulos, start=2):
        c = ws.cell(row=fila, column=col, value=t)
        c.fill = PatternFill("solid", fgColor=COLOR_HEADER)
        c.font = FUENTE_HEADER
        c.alignment = ALIGN_HEADER
        c.border = BORDE_CELDA
    ws.row_dimensions[fila].height = 22
    fila += 1
    # Datos
    for valor, desc in filas_datos:
        c1 = ws.cell(row=fila, column=2, value=valor)
        c1.font = FUENTE_NEGRITA
        c1.alignment = ALIGN_TEXTO
        c1.border = BORDE_CELDA
        c1.fill = PatternFill("solid", fgColor=COLOR_BANDA)

        c2 = ws.cell(row=fila, column=3, value=desc)
        c2.font = FUENTE_NORMAL
        c2.alignment = ALIGN_TEXTO
        c2.border = BORDE_CELDA
        c2.fill = PatternFill("solid", fgColor=COLOR_BANDA)
        ws.row_dimensions[fila].height = 28
        fila += 1
    return fila


# ======================================================================
# HOJA 2: Reportes
# ======================================================================
EJEMPLOS_REPORTES = [
    {
        "ID Reporte": "EJEMPLO 1",
        "Fecha": date(2026, 6, 15),
        "Betatester": "Ana Pérez",
        "Módulo": "Monitoreo en Vivo",
        "Tipo": "Bug funcional",
        "Severidad": "Alta (función clave rota)",
        "Navegador/SO": "Chrome / Windows 11",
        "Pasos para reproducir":
            "1. Iniciar sesión.\n2. Abrir la cámara 'Frontal'.\n"
            "3. Dejar el monitoreo abierto 30 segundos.",
        "Resultado esperado":
            "El stream de video se muestra fluido, sin cortes.",
        "Resultado actual":
            "El video se congela a los 5 segundos y aparece el mensaje "
            "'Reconexión...' indefinidamente.",
        "Evidencia (archivo/ruta)": "captura_monitoreo.png",
        "Estado": "En revisión",
        "Notas/Resolución":
            "⚠️ FILA DE EJEMPLO — bórrala antes de enviar el archivo.",
    },
    {
        "ID Reporte": "EJEMPLO 2",
        "Fecha": date(2026, 6, 15),
        "Betatester": "Ana Pérez",
        "Módulo": "UI/UX",
        "Tipo": "Error visual/CSS",
        "Severidad": "Baja (cosmético/menor)",
        "Navegador/SO": "Firefox / Android",
        "Pasos para reproducir":
            "1. Abrir la app en el móvil.\n2. Ir a la galería de capturas.",
        "Resultado esperado":
            "Los botones de acción se ven alineados dentro de la tarjeta.",
        "Resultado actual":
            "El botón 'Descargar' se sale del borde en pantallas pequeñas.",
        "Evidencia (archivo/ruta)": "",
        "Estado": "Nuevo",
        "Notas/Resolución":
            "⚠️ FILA DE EJEMPLO — bórrala antes de enviar el archivo.",
    },
]


def crear_hoja_reportes(wb: Workbook):
    ws = wb.create_sheet("Reportes")
    ws.sheet_properties.tabColor = "C00000"
    ws.sheet_view.showGridLines = False

    n_col = len(REPORTES_HEADERS)

    # Encabezados en fila 1
    llenar_encabezados(ws, REPORTES_HEADERS, COLOR_HEADER, fila=1)

    # Anchos de columna
    anchos = {
        "A": 12,   # ID Reporte
        "B": 13,   # Fecha
        "C": 20,   # Betatester
        "D": 30,   # Módulo
        "E": 24,   # Tipo
        "F": 24,   # Severidad
        "G": 18,   # Navegador/SO
        "H": 45,   # Pasos
        "I": 45,   # Resultado esperado
        "J": 45,   # Resultado actual
        "K": 24,   # Evidencia
        "L": 15,   # Estado
        "M": 45,   # Notas
    }
    for col, ancho in anchos.items():
        ws.column_dimensions[col].width = ancho

    fila = 2
    # Filas de EJEMPLO (2)
    ejemplo_fill = PatternFill("solid", fgColor=COLOR_EJEMPLO)
    for ej in EJEMPLOS_REPORTES:
        for col_idx, header in enumerate(REPORTES_HEADERS, start=1):
            valor = ej.get(header, "")
            celda = ws.cell(row=fila, column=col_idx, value=valor)
            celda.fill = ejemplo_fill
            celda.border = BORDE_CELDA
            celda.alignment = ALIGN_TEXTO
            if header in ("ID Reporte", "Notas/Resolución"):
                celda.font = FUENTE_EJEMPLO
            else:
                celda.font = FUENTE_NORMAL
            if header == "Fecha" and isinstance(valor, date):
                celda.number_format = "YYYY-MM-DD"
        ws.row_dimensions[fila].height = 60
        fila += 1

    # 50 filas de datos (RPT-001 .. RPT-050)
    for i in range(1, 51):
        ws.cell(row=fila, column=1, value=f"RPT-{i:03d}")  # ID
        ws.cell(row=fila, column=12, value="Nuevo")        # Estado
        for col_idx in range(1, n_col + 1):
            celda = ws.cell(row=fila, column=col_idx)
            celda.border = BORDE_CELDA
            celda.alignment = ALIGN_TEXTO
            celda.font = FUENTE_NORMAL
            if celda.column_letter == "B":
                celda.number_format = "YYYY-MM-DD"
            if (i % 2) == 0:
                celda.fill = PatternFill("solid", fgColor=COLOR_BANDA)
        ws.row_dimensions[fila].height = 42
        fila += 1

    ultima_fila = fila - 1  # 53

    # Validaciones de datos (desplegables) -> referencian la hoja oculta Listas
    # Módulo (D), Tipo (E), Severidad (F), Estado (L)
    anadir_lista(ws, "D", 2, ultima_fila, "Listas", "A", len(MODULOS))
    anadir_lista(ws, "E", 2, ultima_fila, "Listas", "B", len(TIPOS))
    anadir_lista(ws, "F", 2, ultima_fila, "Listas", "C", len(SEVERIDADES))
    anadir_lista(ws, "L", 2, ultima_fila, "Listas", "D", len(ESTADOS))

    # Autofiltro + congelar fila 1
    ws.auto_filter.ref = f"A1:{letra(n_col)}{ultima_fila}"
    ws.freeze_panes = "A2"

    return ws


# ======================================================================
# HOJA 3: Checklist de Pruebas
# ======================================================================
def crear_hoja_checklist(wb: Workbook):
    ws = wb.create_sheet("Checklist de Pruebas")
    ws.sheet_properties.tabColor = "375623"
    ws.sheet_view.showGridLines = False

    llenar_encabezados(ws, CHECKLIST_HEADERS, COLOR_HEADER_CHECK, fila=1)

    anchos = {"A": 6, "B": 58, "C": 14, "D": 20, "E": 22, "F": 42}
    for col, ancho in anchos.items():
        ws.column_dimensions[col].width = ancho

    fila = 2
    for i, func in enumerate(FUNCIONALIDADES, start=1):
        ws.cell(row=fila, column=1, value=i).alignment = ALIGN_CENTRO
        ws.cell(row=fila, column=2, value=func).alignment = ALIGN_TEXTO
        ws.cell(row=fila, column=3).alignment = ALIGN_CENTRO  # ¿Probada?
        ws.cell(row=fila, column=4).alignment = ALIGN_CENTRO  # Resultado
        ws.cell(row=fila, column=5).alignment = ALIGN_TEXTO   # ID relacionado
        ws.cell(row=fila, column=6).alignment = ALIGN_TEXTO   # Comentarios
        for c in range(1, len(CHECKLIST_HEADERS) + 1):
            celda = ws.cell(row=fila, column=c)
            celda.border = BORDE_CELDA
            celda.font = FUENTE_NORMAL
            if (i % 2) == 0:
                celda.fill = PatternFill("solid", fgColor=COLOR_BANDA)
        ws.row_dimensions[fila].height = 26
        fila += 1

    ultima_fila = fila - 1

    # Desplegables: ¿Probada? (C) y Resultado (D)
    anadir_lista(ws, "C", 2, ultima_fila, "Listas", "E", len(OPC_PROBADA))
    anadir_lista(ws, "D", 2, ultima_fila, "Listas", "F", len(OPC_RESULTADO))

    ws.auto_filter.ref = f"A1:{letra(len(CHECKLIST_HEADERS))}{ultima_fila}"
    ws.freeze_panes = "A2"
    return ws


# ======================================================================
# HOJA 4 (oculta): Listas
# ======================================================================
def crear_hoja_listas(wb: Workbook):
    ws = wb.create_sheet("Listas")
    columnas = {
        1: ("Módulo", MODULOS),
        2: ("Tipo", TIPOS),
        3: ("Severidad", SEVERIDADES),
        4: ("Estado", ESTADOS),
        5: ("¿Probada?", OPC_PROBADA),
        6: ("Resultado", OPC_RESULTADO),
    }
    for col, (titulo, valores) in columnas.items():
        ws.cell(row=1, column=col, value=titulo).font = FUENTE_NEGRITA
        for i, v in enumerate(valores, start=1):
            ws.cell(row=i, column=col, value=v)
    # Ocultar la hoja (queda accesible para las validaciones pero no se ve)
    ws.sheet_state = "hidden"
    return ws


# ======================================================================
# Verificación post-escritura
# ======================================================================
def verificar(ruta):
    wb = load_workbook(ruta)
    hojas = wb.sheetnames
    assert "Instrucciones" in hojas, "Falta la hoja 'Instrucciones'"
    assert "Reportes" in hojas, "Falta la hoja 'Reportes'"
    assert "Checklist de Pruebas" in hojas, "Falta la hoja 'Checklist de Pruebas'"
    assert "Listas" in hojas, "Falta la hoja oculta 'Listas'"

    ws_r = wb["Reportes"]
    headers = [ws_r.cell(row=1, column=c).value for c in range(1, 14)]
    assert headers == REPORTES_HEADERS, f"Encabezados Reportes incorrectos: {headers}"
    assert len(ws_r.data_validations.dataValidation) >= 4, \
        "Faltan validaciones en Reportes"

    ws_c = wb["Checklist de Pruebas"]
    headers_c = [ws_c.cell(row=1, column=c).value for c in range(1, 7)]
    assert headers_c == CHECKLIST_HEADERS, f"Encabezados Checklist incorrectos: {headers_c}"
    assert len(ws_c.data_validations.dataValidation) >= 2, \
        "Faltan validaciones en Checklist"

    assert wb["Listas"].sheet_state == "hidden", "La hoja Listas no está oculta"

    return {
        "hojas": hojas,
        "validaciones_reportes": len(ws_r.data_validations.dataValidation),
        "validaciones_checklist": len(ws_c.data_validations.dataValidation),
        "filas_checklist": ws_c.max_row - 1,
    }


# ======================================================================
# Main
# ======================================================================
def main():
    wb = Workbook()
    # Eliminar la hoja por defecto que crea openpyxl
    hoja_defecto = wb.active
    if hoja_defecto is not None:
        wb.remove(hoja_defecto)

    crear_hoja_instrucciones(wb)
    crear_hoja_reportes(wb)
    crear_hoja_checklist(wb)
    crear_hoja_listas(wb)

    # La hoja activa al abrir sea "Instrucciones"
    wb.active = 0

    wb.save(SALIDA)

    info = verificar(SALIDA)
    tam = os.path.getsize(SALIDA)
    print("✅ Ficha generada correctamente:")
    print(f"   Ruta           : {SALIDA}")
    print(f"   Tamaño         : {tam:,} bytes ({tam/1024:.1f} KB)")
    print(f"   Hojas          : {info['hojas']}")
    print(f"   Validaciones Reportes  : {info['validaciones_reportes']}")
    print(f"   Validaciones Checklist : {info['validaciones_checklist']}")
    print(f"   Funcionalidades (Checklist): {info['filas_checklist']}")


if __name__ == "__main__":
    main()
